#!/usr/bin/python
'''
## fortigetX.py ##
Queries a Fortinet Fortigate firewall unit at a given IP address. 
It documents the firewall policies and objects into xlsx spreadsheet.

Feel free to use this code for any purpose, change and sell it for a fortune.
The code is provided "AS IS" and the author is not responsible for any damage done by it. 
(I guess "the code" is too pompous for this ugly snippet anyway)

==== Changes you might wanna make to get it to work====

pxssh method I use to connect to the device requires a recognizable prompt.
Mine looks like "blah blah FW01 #". Change the global PROMPT_REGEXP function if yours does not end with "FW\d+ #"

'''

import re
import sys
import pxssh
import getpass
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Style, Font, Color, PatternFill, Border, Side


ARGV = sys.argv
USAGE = ARGV[0] + ' <Firewall hostname or IP> <sorting criterion>' + '''

sorting criterion: src[intf] | dst[intf] | num[ber]
'''

# prompt for the pxssh to look for
PROMPT_REGEXP = 'FW\d+ # '

# Username to use for login
USERNAME = 'admin_or_whatever_you_have_there'

def FwConnection(ip):
	'''
	Gets a pxssh connection to a Fortigate unit.
	Args:
		ip = ip address of a Fortigate unit. Hostname should do perfectly well too.
	Returns:
		ssh - an established pxssh connection
	'''
	ssh = pxssh.pxssh()
	password = getpass.getpass('password: ')
	username = USERNAME
	# Set the output not to use pager like "more" or "less"
	nomore = 'config system console\n' + 'set output standard\n' + 'end'
	ssh.PROMPT = PROMPT_REGEXP
	try:
		ssh.login(ip, username, password, auto_prompt_reset=False, original_prompt=' # ')
		ssh.sendline(nomore)
		ssh.prompt()
	except pxssh.ExceptionPxssh as e:
		print("Login failed")
		print(e)
	return ssh


def GetSystemInterface(ssh):
	'''
	Gets interfaces statuses and settings.
	Args:
		ssh - established pxssh connection
	Returns:
		port_data = {'ip': <ip address>, 'mask': <mask>, 'status': <link up/down>}
	'''
	ssh.sendline('get system interface')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	port_info = {}

	re_port = re.compile('name:\s+(.+?)\s+(mode:\s+\w*\s*|)ip\:\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\s+status\:\s+(\w+).*')

	for line in raw_out_list:
		port_data_match = re_port.match(line)
		if port_data_match:
			port_info[port_data_match.groups()[0]] = {'ip': port_data_match.groups()[2],
		 			  								  'mask': port_data_match.groups()[3],
		 			  								  'status': port_data_match.groups()[4]}

	return port_info


def GetPolicies(ssh):
	'''
	Gets policies for a given port.
	Args:
		ssh - established pxssh connection
		port - interface name 
	Returns:
		policies = {'number': <policy id>,
				    'srcintf': [<source interface>, ...], 
				    'dstintf': [<destination interface>, ...],
				    'srcaddr': [<source ip>, ...],
				    'dstaddr': [<destination ip>, ...],
				    'service': [<service>, ...],
				    'action': <action>,
				    'nat': bool,
				    'natpool': <natpool_name}
	'''
	policies = []
	ssh.sendline('show firewall policy')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit (\d{1,3}).*')
	re_next = re.compile('\s+next')
	re_dstintf = re.compile('\s+set dstintf \"(.+)\"')
	re_srcintf = re.compile('\s+set srcintf \"(.+)\"')
	re_dstaddr = re.compile('\s+set dstaddr \"(.+)\"')
	re_srcaddr = re.compile('\s+set srcaddr \"(.+)\"')
	re_service = re.compile('\s+set service \"(.+)\"')
	re_action = re.compile('\s+set action (.+)\r.*')
	re_nat = re.compile('\s+set nat enable')
	re_natpool = re.compile('\s+set poolname "(.+)"')
	re_comments = re.compile('\s+set comments \"(.+)\"')

	policy = {}
	for line in raw_out_list:
		match_next = re_next.match(line)
		if match_next and policy:
			# If SSL-VPN - there is no service
			if policy['action'] == 'ssl-vpn':
				policy['service'] = ['n/a']
			if 'comments' not in policy:
				policy['comments'] = ''

			policies.append(policy)
			policy = {}
			continue
		else:
			edit_match = re_edit.match(line)
			if edit_match:
				policy['number'] = int(edit_match.groups()[0])
				# Se NAT to false until we prove opposite
				policy['nat'] = False
			
			srcintf_match = re_srcintf.match(line)
			if srcintf_match:
				policy['srcintf'] = srcintf_match.groups()[0].split('\" \"')

			dstintf_match = re_dstintf.match(line)
			if dstintf_match:
				policy['dstintf'] = dstintf_match.groups()[0].split('\" \"')

			srcaddr_match = re_srcaddr.match(line)
			if srcaddr_match:
				policy['srcaddr'] = srcaddr_match.groups()[0].split('\" \"')

			dstaddr_match = re_dstaddr.match(line)
			if dstaddr_match:
				policy['dstaddr'] = dstaddr_match.groups()[0].split('\" \"')

			service_match = re_service.match(line)
			if service_match:
				policy['service'] = service_match.groups()[0].split('\" \"')

			action_match = re_action.match(line)
			if action_match:
				policy['action'] = action_match.groups()[0]

			nat_match = re_nat.match(line)
			if nat_match:
				policy['nat'] = True

			natpool_match = re_natpool.match(line)
			if natpool_match:
				policy['natpool'] = natpool_match.groups()[0]

			comments_match = re_comments.match(line)
			if comments_match:
				policy['comments'] = comments_match.groups()[0]

	return policies


def SortPolicies(policies, key='srcintf'):
	'''
	Converts list of dicts [{},{}] to dict of dicts {key: {}, key: {}}
	so it can be sorted by key.
	Args:
		policies - [{}. {}]
		key - srcintf | dstints | number
	Returns:
		policies - [key: {}, key: {}]
	'''
	policies_sorted = {}
	if key not in ['srcintf', 'dstintf', 'number']:
		print 'SortPolicies: wrong key' + key
		raise SystemExit(1)


	for policy in policies:
		if key in ['srcintf', 'dstintf']:
			if key == 'srcintf':
				skey = 'dstintf'
			else:
				skey = 'srcintf'
			# {src|dst}intf is a list. Most likely it has just one member, but we have to think of the worst
			# since interfaces are not unique in policies we have to use the policy number as well
			# not very neat sorting, but representative anyway. 
			policies_sorted[','.join(sorted(policy[key]) + sorted(policy[skey])) + '--' + str(policy['number'])] = policy
		else:
			policies_sorted[policy[key]] = policy
	return policies_sorted


def GetAddrGroups(ssh):
	'''
	Gets all the Address Group objects and their members.
	Args:
		ssh - established pxssh connection
	Returns:
		grp_members = {Grp_obj: [addr_obj1, ...]}
	'''
	ssh.sendline('show firewall addrgrp')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_members = re.compile('\s+set member "(.+)"')

	# Build full dict of groups {grp: [addr1, ...]}
	grp_members = {}

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		members_match = re_members.match(line)

		if edit_match:
			curr_grp = edit_match.groups()[0]
		if members_match:
			grp_members[curr_grp] = members_match.groups()[0].split('" "')

	return grp_members


def GetVIPGroups(ssh):
	'''
	Gets all the VIP Group objects and their members.
	Args:
		ssh - established pxssh connection
	Returns:
		grp_members = {Grp_obj: [vip_obj1, ...]}
	'''
	ssh.sendline('show firewall vipgrp')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_members = re.compile('\s+set member "(.+)"')

	# Build full dict of groups {grp: [addr1, ...]}
	grp_members = {}

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		members_match = re_members.match(line)

		if edit_match:
			curr_grp = edit_match.groups()[0]
		if members_match:
			grp_members[curr_grp] = members_match.groups()[0].split('" "')

	return grp_members


def GetServiceGroups(ssh):
	'''
	Gets all the Address Group objects and their members.
	Args:
		ssh - established pxssh connection
	Returns:
		grp_members = {Srv_GRP_obj: [srv_obj1, ...]}
	'''
	ssh.sendline('show firewall service group' )
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_members = re.compile('\s+set member "(.+)"')

	# Build full dict of groups {grp: [addr1, ...]}
	grp_members = {}

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		members_match = re_members.match(line)

		if edit_match:
			curr_grp = edit_match.groups()[0]
		if members_match:
			grp_members[curr_grp] = members_match.groups()[0].split('" "')

	return grp_members


def GetServices(ssh):	

	ssh.sendline('config firewall service custom\nshow full-configuration\nend' )
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_tcp = re.compile('\s+set tcp-portrange (\d.*\d).*')
	re_udp = re.compile('\s+set udp-portrange (\d.*\d).*')

	srvs = {}

	tcp = ''
	udp = ''

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		tcp_match = re_tcp.match(line)
		udp_match = re_udp.match(line)

		if edit_match:
			# We hit a new service clause - let's add the current one
			if tcp and udp:
				srvs[curr_srv] = tcp + ' ' + udp
			elif tcp:
				srvs[curr_srv] = tcp
			elif udp:
				srvs[curr_srv] = udp

			tcp = ''
			udp = ''

			curr_srv = edit_match.groups()[0]

		if tcp_match:
			tcp = 'TCP/' + tcp_match.groups()[0]
		if udp_match:
			udp = 'UDP/' + udp_match.groups()[0]
	# This is the last found. I know it's ugly, but I'm tired and want to go home
	if tcp and udp:
		srvs[curr_srv] = tcp + ' ' + udp
	elif tcp:
		srvs[curr_srv] = tcp
	elif udp:
		srvs[curr_srv] = udp

	return srvs


def dec2cidr(dec_netmask):
	'''
	Converts dot decimal netmask notation to cidr
	Args:
		dec_netmask - dot decimal netmask like 255.255.255.0
	Returns:
		cidr, e.g 24
	'''
	cidr = 0
	octet_len = {'0': 0,
				 '128': 1,
				 '192': 2,
				 '224': 3,
				 '240': 4,
				 '248': 5,
				 '252': 6,
				 '254': 7,
				 '255': 8}

	octets = dec_netmask.split('.')
	for octet in octets:
		try:
			cidr += octet_len[octet]
		except KeyError:
			print 'Weird mask ' + dec_netmask
	
	return cidr


def GetIPs(ssh):
	ips = {}

	ssh.sendline('show firewall address')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_next = re.compile('\s+next')
	re_range = re.compile('\s+set type iprange.*')
	re_end_ip = re.compile('\s+set end-ip (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}).*')
	re_start_ip = re.compile('\s+set start-ip (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}).*')
	re_subnet = re.compile('\s+set subnet (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}) (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}).*')

	# Flag indicating that this object is an ip range rather then subnet
	is_range = 0
	ip = ''
	start_ip = ''
	end_ip = ''
	curr_ip = ''

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		range_match = re_range.match(line)
		start_ip_match = re_start_ip.match(line)
		end_ip_match = re_end_ip.match(line)
		subnet_match = re_subnet.match(line)

		if edit_match:
			curr_ip = edit_match.groups()[0]

		if range_match:
			is_range = 1
		if start_ip_match and is_range:
			start_ip = start_ip_match.groups()[0]
		if end_ip_match and is_range:
			end_ip = end_ip_match.groups()[0]
		if subnet_match:
			ip = subnet_match.groups()[0] + '/' + str(dec2cidr(subnet_match.groups()[1]))
		if is_range and start_ip and end_ip:
			ip = start_ip + '-' + end_ip
		if ip:
			ips[curr_ip] = ip
			is_range = 0
			start_ip = ''
			end_ip = ''
			ip = ''
	return ips


def GetVIPs(ssh):

	vips = {}

	ssh.sendline('show firewall vip')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_next = re.compile('\s+next')
	re_extip = re.compile('\s+set extip (.+\d).*')
	re_mappedip = re.compile('\s+set mappedip "(.+\d)".*')
	re_portfw = re.compile('\s+set portforward enable')
	re_proto = re.compile('\s+set protocol (\w+)')
	re_extport = re.compile('\s+set extport (\d{1,5}-?\d{0,5})')
	re_mappedport = re.compile('\s+set mappedport (\d{1,5}-?\d{0,5})')

	vip = ''
	ext_ip = ''
	mapped_ip = ''
	curr_ip = ''
	extport = ''
	mappedport = ''
	proto = 'tcp'
	portfw = False

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		next_match = re_next.match(line)
		extip_match = re_extip.match(line)
		mappedip_match = re_mappedip.match(line)
		portfw_match = re_portfw.match(line)
		proto_match = re_proto.match(line)
		extport_match = re_extport.match(line)
		mappedport_match = re_mappedport.match(line)

		if next_match:
			vip += ext_ip + '->' + mapped_ip
			if portfw:
				vip += '(' + proto + ':' + extport + '->' + mappedport + ')'
			vips[curr_ip] = vip
			ext_ip = ''
			mapped_ip = ''
			vip = ''
			proto = 'tcp'
			extport = ''
			mappedport = ''
			portfw = False		

		if edit_match:
			curr_ip = edit_match.groups()[0]

		if extip_match:
			ext_ip = extip_match.groups()[0]

		if mappedip_match:
			mapped_ip = mappedip_match.groups()[0]

		if portfw_match:
			portfw = True

		if proto_match:
			proto = proto_match.groups()[0]

		if extport_match:
			extport = extport_match.groups()[0]

		if mappedport_match:
			mappedport = mappedport_match.groups()[0]			

	return vips


def GetIpPools(ssh):
	'''
	Gets full list of al registered IP pools with there names, start and end ip.
	Args:
		ssh - established pxssh connection
	Returns:
		ip_pools = {pool_name: [start_ip, end_ip]}
	'''
	ssh.sendline('show firewall ippool')
	ssh.prompt()
	raw_out = ssh.before

	raw_out_list = raw_out.split('\n')

	re_edit = re.compile('\s+edit "(.+)"')
	re_startip = re.compile('\s+set startip (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})')
	re_endip = re.compile('\s+set endip (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})')

	ip_pools = {}

	for line in raw_out_list:
		edit_match = re_edit.match(line)
		startip_match = re_startip.match(line)
		endip_match = re_endip.match(line)

		if edit_match:
			curr_pool = edit_match.groups()[0]
			ip_pools[curr_pool] = {}
		if startip_match:
			ip_pools[curr_pool]['start'] = startip_match.groups()[0]
		if endip_match:
			ip_pools[curr_pool]['end'] = endip_match.groups()[0]

	return ip_pools	


def ParseArgs(ARGV):
	'''
	'''
	if len(ARGV) != 3:
		print USAGE
		raise SystemExit(1)

	params = ' '.join(ARGV[1:])
	re_param = re.compile('(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}) .+')
	param_match = re_param.match(params)

	# if no ip address followed by some text found
	if (not param_match) or (ARGV[1] == 'help'):
		print USAGE
		raise SystemExit(0)

	sorting = {'src': 'srcintf',
			   'dst': 'dstintf',
			   'srcintf': 'srcintf',
			   'dstintf': 'dstintf',
			   'num': 'number',
			   'number': 'number'}

 	if ARGV[2] in sorting:
		sort_key = sorting[ARGV[2]]
	else:
		print USAGE
		raise SystemExit(0)

	return sort_key


def main(ARGV):

	sort_key = ParseArgs(ARGV)

	ip = ARGV[1]
	ssh = FwConnection(ip)
	policies = GetPolicies(ssh)
	ips = GetIPs(ssh)
	srvs = GetServices(ssh)
	vips = GetVIPs(ssh)
	ip_grps = GetAddrGroups(ssh)
	srv_grps = GetServiceGroups(ssh)
	vip_grps = GetVIPGroups(ssh)
	ip_pools = GetIpPools(ssh)
	port_info = GetSystemInterface(ssh)
	ssh.logout()

	policies_sorted = SortPolicies(policies, sort_key)

	wb = Workbook(guess_types=False)
	wb.remove_sheet(wb.active)
	ws_pol = wb.create_sheet(title='policies')
	ws_if = wb.create_sheet(title='interfaces')
	ws_ip = wb.create_sheet(title='addresses')
	ws_ip_grp = wb.create_sheet(title='addr groups')
	ws_srv = wb.create_sheet(title='services')
	ws_srv_grp = wb.create_sheet(title='srv groups')
	ws_vip = wb.create_sheet(title='VIPs')
	ws_vip_grp = wb.create_sheet(title='VIP groups')
	ws_ip_pools = wb.create_sheet(title='IP pools')

	style_border_thin_all = Border(left=Side(style='thin'), 
                    			   right=Side(style='thin'), 
                    			   top=Side(style='thin'), 
                    			   bottom=Side(style='thin'))
	style_border_thin_tb = Border(top=Side(style='thin'), 
                    			  bottom=Side(style='thin'))
	style_border_fat_r = Border(right=Side(style='thick'))

	style_fill_green = PatternFill(patternType='solid', fgColor=Color('66CC99'))
	style_fill_gray = PatternFill(patternType='solid', fgColor=Color('D0D0D0'))
	style_fill_cyan = PatternFill(patternType='solid', fgColor=Color('CCCCFF'))
	style_fill_yellow = PatternFill(patternType='solid', fgColor=Color('FFFFCC'))
	style_fill_lgreen = PatternFill(patternType='solid', fgColor=Color('64FE2E'))
	style_fill_dred = PatternFill(patternType='solid', fgColor=Color('B40404'))
	style_fill_red = PatternFill(patternType='solid', fgColor=Color('FF0000'))
	style_fill_blue = PatternFill(patternType='solid', fgColor=Color('A9D0F5'))
	style_fill_lblue = PatternFill(patternType='solid', fgColor=Color('EFF5FB'))
	
	style_font_bold = Font(bold=True)
	style_font_blue = Font(color='0000FF')
	style_font_dred = Font(color='B40404')
	style_font_lgreen = Font(color='64FE2E')

	style_caption = Style(font=style_font_bold, 
						  border=style_border_thin_all, 
						  fill=style_fill_gray)
	style_group = Style(font=style_font_bold, 
						border=style_border_thin_tb, 
						fill=style_fill_yellow)
	style_policy = Style(font=style_font_bold, 
						 border=style_border_thin_all, 
						 fill=style_fill_green)
	style_srvgroup = Style(font=style_font_bold, 
						   border=style_border_thin_tb, 
						   fill=style_fill_cyan)
	style_vipgroup = Style(font=style_font_bold, 
						   border=style_border_thin_tb, 
						   fill=style_fill_blue)
	style_vip = Style(fill=style_fill_lblue)

	ws_pol['A1'] = ' num'
	ws_pol['B1'] = ' SOURCE'
	ws_pol['C1'] = ' IP'
	ws_pol['D1'] = ' DESTINATION'
	ws_pol['E1'] = ' IP'
	ws_pol['F1'] = ' Service'
	ws_pol['G1'] = ' Port'
	ws_pol['H1'] = ' NAT'
	ws_pol['I1'] = ' Comments'

	ws_pol.column_dimensions['A'].width = 5
	ws_pol.column_dimensions['B'].width = 30
	ws_pol.column_dimensions['C'].width = 35
	ws_pol.column_dimensions['C'].outline_level = 1
	ws_pol.column_dimensions['D'].width = 30
	ws_pol.column_dimensions['E'].width = 48
	ws_pol.column_dimensions['E'].outline_level = 1
	ws_pol.column_dimensions['F'].width = 20
	ws_pol.column_dimensions['G'].width = 20
	ws_pol.column_dimensions['G'].outline_level = 1
	ws_pol.column_dimensions['H'].width = 55
	ws_pol.column_dimensions['I'].width = 100

	ws_if['A1'] = 'Interface'
	ws_if['B1'] = 'IP'
	ws_if['C1'] = 'Mask'
	ws_if['D1'] = 'Status'

	ws_if.column_dimensions['A'].width = 20
	ws_if.column_dimensions['B'].width = 30
	ws_if.column_dimensions['C'].width = 30
	ws_if.column_dimensions['D'].width = 10

	ws_ip['A1'] = 'Object name'
	ws_ip['B1'] = 'address'

	ws_ip.column_dimensions['A'].width = 35
	ws_ip.column_dimensions['B'].width = 35

	ws_srv['A1'] = 'Object name'
	ws_srv['B1'] = 'services'

	ws_srv.column_dimensions['A'].width = 35
	ws_srv.column_dimensions['B'].width = 35

	ws_vip['A1'] = 'Object name'
	ws_vip['B1'] = 'VIP mapping'

	ws_vip.column_dimensions['A'].width = 35
	ws_vip.column_dimensions['B'].width = 48

	ws_ip_grp['A1'] = 'Group name'
	ws_ip_grp['B1'] = 'IP object name'
	ws_ip_grp['C1'] = 'address'

	ws_ip_grp.column_dimensions['A'].width = 25
	ws_ip_grp.column_dimensions['B'].width = 30
	ws_ip_grp.column_dimensions['C'].width = 20
	ws_ip_grp.column_dimensions['C'].outline_level = 1

	ws_srv_grp['A1'] = 'Group name'
	ws_srv_grp['B1'] = 'SRV object name'
	ws_srv_grp['C1'] = 'services'

	ws_srv_grp.column_dimensions['A'].width = 25
	ws_srv_grp.column_dimensions['B'].width = 30
	ws_srv_grp.column_dimensions['C'].width = 20
	ws_srv_grp.column_dimensions['C'].outline_level = 1

	ws_vip_grp['A1'] = 'Group name'
	ws_vip_grp['B1'] = 'VIP object name'
	ws_vip_grp['C1'] = 'address'

	ws_vip_grp.column_dimensions['A'].width = 25
	ws_vip_grp.column_dimensions['B'].width = 30
	ws_vip_grp.column_dimensions['C'].width = 35
	ws_vip_grp.column_dimensions['C'].outline_level = 1

	ws_ip_pools['A1'] = 'Pool name'
	ws_ip_pools['B1'] = 'address range'

	ws_ip_pools.column_dimensions['A'].width = 35
	ws_ip_pools.column_dimensions['B'].width = 35

	for ws in wb.worksheets:
		for col in ws.columns: # col[0] returns contains first cell e.g. 'A1', 'B1' ...
			if col:
				col[0].style = style_caption


	# Writing addresses sheet
	i = 2
	for addr in sorted(ips):
		ws_ip['A' + str(i)] = addr
		ws_ip['B' + str(i)] = ips[addr]
		i += 1

	# Writing interfaces sheet
	i = 2
	for intf in sorted(port_info):
		ws_if['A' + str(i)] = intf
		ws_if['B' + str(i)] = port_info[intf]['ip']
		ws_if['C' + str(i)] = port_info[intf]['mask']
		ws_if['D' + str(i)] = port_info[intf]['status']
		i += 1

	# Writing Services sheet
	i = 2
	for srv in sorted(srvs):
		ws_srv['A' + str(i)] = srv
		ws_srv['B' + str(i)] = srvs[srv]
		i += 1

	# Writing VIPs sheet
	i = 2
	for vip in sorted(vips):
		ws_vip['A' + str(i)] = vip
		ws_vip['B' + str(i)] = vips[vip]
		i += 1

	# Writing IPpools sheet
	i = 2
	for pool in sorted(ip_pools):
		ws_ip_pools['A' + str(i)] = pool
		ws_ip_pools['B' + str(i)] = ip_pools[pool]['start'] + '-' + ip_pools[pool]['end']
		i += 1

	# Writing IP groups sheet
	i = 2
	last_ip_num = str(len(ips) + 1)

	for grp in sorted(ip_grps):
		ws_ip_grp['A' + str(i)] = grp

		# Put a list of member IPs in next column.
		#
		for row in ws_ip_grp['A' + str(i):'C' + str(i)]:
			for cell in row: 
				cell.style = style_group
		i += 1
		for addr in ip_grps[grp]:
			ws_ip_grp['B' + str(i)] = addr
			vlookup = '=VLOOKUP(B' + str(i) + ',' + ws_ip.title + '!A2:B' + last_ip_num + ',2,FALSE())'
			ws_ip_grp['C' + str(i)]= vlookup
			i += 1

	# Writing srv groups sheet
	i = 2
	last_srv_num = str(len(srvs) + 1)

	for grp in sorted(srv_grps):
		ws_srv_grp['A' + str(i)] = grp

		# Put a list of member SRVs in next column.
		#
		for row in ws_srv_grp['A' + str(i):'C' + str(i)]:
			for cell in row: 
				cell.style = style_srvgroup
		i += 1
		for addr in srv_grps[grp]:
			ws_srv_grp['B' + str(i)] = addr
			vlookup = '=VLOOKUP(B' + str(i) + ',' + ws_srv.title + '!A2:B' + last_srv_num + ',2,FALSE())'
			ws_srv_grp['C' + str(i)]= vlookup
			i += 1

	# Writing VIP groups sheet
	i = 2
	last_vip_num = str(len(vips) + 1)

	for grp in sorted(vip_grps):
		ws_vip_grp['A' + str(i)] = grp

		# Put a list of member IPs in next column.
		#
		for row in ws_vip_grp['A' + str(i):'C' + str(i)]:
			for cell in row: 
				cell.style = style_vipgroup
		i += 1
		for addr in vip_grps[grp]:
			ws_vip_grp['B' + str(i)] = addr
			vlookup = '=VLOOKUP(B' + str(i) + ',' + ws_vip.title + '!A2:B' + last_vip_num + ',2,FALSE())'
			ws_vip_grp['C' + str(i)]= vlookup
			i += 1


	# Our main thing

	i = 2 # row index for 

	for policy in sorted(policies_sorted):
		max_row_depth = 2 # next policy item should start after this row
		ws_pol['A' + str(i)] = policies_sorted[policy]['number']

		src_if_list = []
		dst_if_list = []

		for intf in sorted(policies_sorted[policy]['srcintf']):
			src_if_list.append(intf + ' (' + port_info[intf]['ip'] + '/' + str(dec2cidr(port_info[intf]['mask'])) + ' ' + port_info[intf]['status'] + ')')

		for intf in sorted(policies_sorted[policy]['dstintf']):
			dst_if_list.append(intf + ' (' + port_info[intf]['ip'] + '/' + str(dec2cidr(port_info[intf]['mask'])) + ' ' + port_info[intf]['status'] + ')')

		ws_pol['B' + str(i)] = ','.join(src_if_list)
		ws_pol['D' + str(i)] = ','.join(dst_if_list)

		ws_pol['I' + str(i)] = policies_sorted[policy]['comments']

		nat_info = ''
		if policies_sorted[policy]['nat']:
			if 'natpool' in policies_sorted[policy]:
				nat_pool = policies_sorted[policy]['natpool']
				nat_info = nat_pool + ': ' + ip_pools[nat_pool]['start'] + '-' + ip_pools[nat_pool]['end']
			else:
				nat_info = 'NAT to dstintf addr'
		ws_pol['H' + str(i)] = nat_info

		# some styling
		for row in ws_pol['A' + str(i):'H' + str(i)]:
			for cell in row: 
				cell.style = style_policy
		##

		i += 1

		# SRC side
		j = i
		for addr in sorted(policies_sorted[policy]['srcaddr']):
			if addr in ip_grps:
				ws_pol['B' + str(j)] = addr
				
				# some styling for IP group
				for row in ws_pol['B' + str(j):'C' + str(j)]:
					for cell in row: 
						cell.style = style_group
				##
				j += 1
				for ip in sorted(ip_grps[addr]):
					ws_pol['B' + str(j)] = ip
					vlookup = '=VLOOKUP(B' + str(j) + ',' + ws_ip.title + '!A2:B' + last_ip_num + ',2,FALSE())'
					ws_pol['C' + str(j)] = vlookup
					j += 1 
			elif addr in vip_grps:
				ws_pol['B' + str(j)] = addr
				
				# some styling for VIP group
				for row in ws_pol['B' + str(j):'C' + str(j)]:
					for cell in row: 
						cell.style = style_vipgroup
				##
				j += 1
				for vip in sorted(vip_grps[addr]):
					ws_pol['B' + str(j)] = vip
					ws_pol['B' + str(j)].font = style_font_blue
					vlookup = '=VLOOKUP(B' + str(j) + ',' + ws_vip.title + '!A2:B' + last_vip_num + ',2,FALSE())'
					ws_pol['C' + str(j)] = vlookup
					j += 1 
			else:
				ws_pol['B' + str(j)] = addr
				if addr in vips:
					vlookup = '=VLOOKUP(B' + str(j) + ',' + ws_vip.title + '!A2:B' + last_vip_num + ',2,FALSE())'
					ws_pol['B' + str(j)].font = style_font_blue
				else:
					vlookup = '=VLOOKUP(B' + str(j) + ',' + ws_ip.title + '!A2:B' + last_ip_num + ',2,FALSE())'
				ws_pol['C' + str(j)] = vlookup
				j += 1
			j += 1
		max_row_depth = j

		# DST side
		j = i
		for addr in sorted(policies_sorted[policy]['dstaddr']):
			if addr in ip_grps:
				ws_pol['D' + str(j)] = addr
				
				# some styling
				for row in ws_pol['D' + str(j):'E' + str(j)]:
					for cell in row: 
						cell.style = style_group
				##
				j += 1
				for ip in sorted(ip_grps[addr]):
					ws_pol['D' + str(j)] = ip
					vlookup = '=VLOOKUP(D' + str(j) + ',' + ws_ip.title + '!A2:B' + last_ip_num + ',2,FALSE())'
					ws_pol['E' + str(j)] = vlookup
					j += 1 
			elif addr in vip_grps:
				ws_pol['D' + str(j)] = addr
				
				# some styling for VIP group
				for row in ws_pol['D' + str(j):'E' + str(j)]:
					for cell in row: 
						cell.style = style_vipgroup
				##
				j += 1
				for vip in sorted(vip_grps[addr]):
					ws_pol['D' + str(j)] = vip
					ws_pol['D' + str(j)].font = style_font_blue
					vlookup = '=VLOOKUP(D' + str(j) + ',' + ws_vip.title + '!A2:B' + last_vip_num + ',2,FALSE())'
					ws_pol['E' + str(j)] = vlookup
					j += 1 

			else:
				ws_pol['D' + str(j)] = addr
				if addr in vips:
					vlookup = '=VLOOKUP(D' + str(j) + ',' + ws_vip.title + '!A2:B' + last_vip_num + ',2,FALSE())'
					ws_pol['D' + str(j)].font = style_font_blue
				else:
					vlookup = '=VLOOKUP(D' + str(j) + ',' + ws_ip.title + '!A2:B' + last_ip_num + ',2,FALSE())'
				ws_pol['E' + str(j)] = vlookup
				j += 1
			j += 1
		
		if j > max_row_depth:
			max_row_depth = j

		# SRV columns
		j = i
		for srv_obj in sorted(policies_sorted[policy]['service']):
			if srv_obj in srv_grps:
				ws_pol['F' + str(j)] = srv_obj

				# some styling
				for row in ws_pol['F' + str(j):'G' + str(j)]:
					for cell in row: 
						cell.style = style_srvgroup
				##
				j += 1
				for srv in sorted(srv_grps[srv_obj]):
					ws_pol['F' + str(j)] = srv
					vlookup = '=VLOOKUP(F' + str(j) + ',' + ws_srv.title + '!A2:B' + last_srv_num + ',2,FALSE())'
					ws_pol['G' + str(j)] = vlookup
					j += 1 
			else:
				ws_pol['F' + str(j)] = srv_obj
				vlookup = '=VLOOKUP(F' + str(j) + ',' + ws_srv.title + '!A2:B' + last_srv_num + ',2,FALSE())'
				ws_pol['G' + str(j)] = vlookup
				j += 1
			j += 1 

		if j > max_row_depth:
			max_row_depth = j

		i = max_row_depth

	for r in range(1, i - 1):
		ws_pol['B' + str(r)].border = ws_pol['B' + str(r)].border.copy(left=Side(style='thick'))
		ws_pol['D' + str(r)].border = ws_pol['D' + str(r)].border.copy(left=Side(style='thick'))
		ws_pol['F' + str(r)].border = ws_pol['F' + str(r)].border.copy(left=Side(style='thick'))
		ws_pol['I' + str(r)].border = ws_pol['I' + str(r)].border.copy(left=Side(style='thick'))


	wb.save('fwdump.xlsx')


if __name__ == '__main__':
    main(ARGV)
